import numpy as np
from scipy.sparse import (
    eye, kron, diags, csr_matrix, lil_matrix, tril, triu, hstack
)
from numpy.linalg import norm
import time
import scipy.sparse.linalg as spla

def time_it(func):
    """A decorator to measure the execution time of a function and return it."""
    def wrapper(*args, **kwargs):
        # print(f"Executing '{func.__name__}' with h = {h}...")
        start_time = time.time()  # Record start time
        result = func(*args, **kwargs)  # Call the original function
        end_time = time.time()  # Record end time
        execution_time = end_time - start_time
        print(f"Function '{func.__name__}' with h = {h} executed in {execution_time:.4f} seconds.")
        return result  # Return both result and time
    return wrapper

def u0(x,y,z,d):
    """
    function for the boundary conditions
    """

    if d == 2:
        return np.sin(x*y)
    return np.sin(x*y*z)

def f(x,y,z,d):
    """
    function for the internal grid points where 
    -nabla u = f
    """
    if d == 2:
        f = x**2*np.sin(x*y)+y**2*np.sin(x*y)
        return f
    return x**2*np.sin(x*y*z)+y**2*np.sin(x*y*z)+z**2*np.sin(x*y*z)

def create_Ah(
        h: float,
        d: int,
        N: int
):
    """
    Creates the Ah for the poisson equation in d-dimensions
    
    """

    #first create the 1D poisson equation
    Ah_oneD = np.zeros([N,N])
    Ah_oneD[0,0], Ah_oneD[-1,-1] = 2*[h**2]
    for i in range(1,N-1):
        Ah_oneD[i,i] = 2
    for i in range(2,N-1):
        Ah_oneD[i-1,i] = -1
        Ah_oneD[i,i-1] = -1

    if d == 1:
        return Ah_oneD/(h**2)
        
    # The n-dimensional case can be constructed by the sum of kronicker products with identities
    def kron_n(d, j):
        """
        Creates the kronicker product IxIxAhxI
        with d terms and Ah on the j-th place
        """
        I = np.eye(N)
        tempAh = d*[I]
        tempAh[j] = Ah_oneD
        for i in range(d-1):
            tempAh[i+1] = np.kron(tempAh[i],tempAh[i+1])
        return tempAh[-1]
    
    Ah = sum([kron_n(d,k) for k in range(d)])

    # Only the border points will be wrong. 
    # They can be selected by looking at the rows with diagonal not equal to 2*d
    for i in range(N**d):
        if Ah[i,i] != 2*d:
            Ah[i] = np.zeros(N**d)
            Ah[i, i] = h**2

    return Ah/(h**2)

def create_Ah_sparse(
        h: float,
        d: int,
        N: int
):
    """
    Creates the Ah for the Poisson equation in d-dimensions.
    Returns a sparse matrix for efficiency.
    """
    # First create the 1D Poisson equation matrix as a sparse matrix
    diagonals = [2 * np.ones(N), -1 * np.ones(N-1), -1 * np.ones(N-1)]
    offsets = [0, -1, 1]
    Ah_oneD = diags(diagonals, offsets, shape=(N, N), format="csr")

    # Remove the references to boundary points
    Ah_oneD[0,1], Ah_oneD[1,0], Ah_oneD[-1,-2], Ah_oneD[-2,-1] = 0,0,0,0

    if d == 1:
        return Ah_oneD
    
    # The n-dimensional case can be constructed by the sum of Kronecker products with identities
    def kron_n(d, j):
        """
        Creates the Kronecker product I ⊗ I ⊗ Ah ⊗ I
        with d terms and Ah on the j-th place.
        """
        I = eye(N, format="csr")
        tempAh = [I] * d
        tempAh[j] = Ah_oneD
        result = tempAh[0]
        for i in range(1, d):
            result = kron(result, tempAh[i], format="csr")
        return result
    
    Ah = sum(kron_n(d, k) for k in range(d))

    # Convert Ah to LIL format for efficient row modifications
    Ah = Ah.tolil()

    # Fix rows where the diagonal is not equal to 2 * d
    for i in range(N**d):
        r, c, l = row_col_dim(i, N)

        # Check if r coinsides with a border point
        if (r in [0,N-1]) or (c in [0,N-1]) or (l in [0,N-1] and d == 3):
            Ah.rows[i] = [i]  # Set row to only include the diagonal
            Ah.data[i] = [h**2]

    # Convert back to CSR format for efficiency
    Ah = Ah.tocsr()

    return Ah / (h ** 2)

def row_col_dim(index, N):
    """
    Usefull function to get the row, col, (and for 3d dimension) 
    corresponding to an index of f or Ah
    """

    row = index % N
    col = (index // N) % N
    dim = index // N**2
    return row, col, dim

def create_fh(
        h: float,
        d: int,
        N: int
):
    """
    Function that creates fh
    """

    fh = np.zeros(N**d)
    for i in range(N**d):
        # Find the row column and layer corresponding to the index
        # Also transform them to the corresponding x, y (and z) coordinates
        r, c, l = row_col_dim(i, N)
        x, y, z = h*r, h*c, h*l

        # Check if r coinsides with a border point
        if (r in [0,N-1]) or (c in [0,N-1]) or (l in [0,N-1] and d == 3):
            fh[i] = u0(x, y, z, d)
        else:
            # For interior points fill in f()
            fh[i] = f(x, y, z, d)

            # If bordering a boundary we fill add the boundary condition
            if r == 1:
                fh[i] += u0(0, y, z, d)/h**2
            if r == N-2:
                fh[i] += u0(1, y, z, d)/h**2
            if c == 1:
                fh[i] += u0(x, 0, z, d)/h**2
            if c == N-2:
                fh[i] += u0(x, 1, z, d)/h**2
            if l == 1 and d == 3:
                fh[i] += u0(x, y, 0, d)/h**2
            if l == N-2 and d == 3:
                fh[i] += u0(x, y, 1, d)/h**2

    return fh

@time_it
def cholesky_banded(A, bandwidth):

    n = A.shape[0]
    C = np.zeros_like(A)

    for i in range(n):
        # Diagonal entry computation
        sum_diagonal = sum(C[i, k] ** 2 for k in range(max(0, i - bandwidth), i))
        C[i, i] = np.sqrt(A[i, i] - sum_diagonal)

        # Off-diagonal entries within bandwidth
        for j in range(i + 1, min(i + bandwidth + 1, n)):
            sum_off_diagonal = sum(C[i, k] * C[j, k] for k in range(max(0, j - bandwidth), i))
            C[j, i] = (A[j, i] - sum_off_diagonal) / C[i, i]
    
    return C

@time_it
def cholesky_banded_sparse(A, bandwidth):
    """
    Perform Cholesky decomposition on a sparse banded matrix A.
    """
    if not isinstance(A, csr_matrix):
        raise ValueError("Input matrix A must be a scipy.sparse.csr_matrix")

    n = A.shape[0]
    C = lil_matrix(A.shape)  # Use LIL format for efficient row-wise updates

    for i in range(n):
        # Diagonal entry computation
        sum_diagonal = sum(C[i, k] ** 2 for k in range(max(0, i - bandwidth), i))
        C[i, i] = np.sqrt(A[i, i] - sum_diagonal)

        # Off-diagonal entries within bandwidth
        for j in range(i + 1, min(i + bandwidth + 1, n)):
            sum_off_diagonal = sum(C[i, k] * C[j, k] for k in range(max(0, j - bandwidth), i))
            C[j, i] = (A[j, i] - sum_off_diagonal) / C[i, i]

    # # Extract the LU decomposition
    # lu = spla.splu(A)

    # # Extract the lower triangular matrix L from LU decomposition
    # L = lu.L  # L is a sparse lower triangular matrix
    # U = lu.U

    return C  # Convert to CSR format for efficient storage and operations

@time_it
def sparse_lu(A):

    # Extract the LU decomposition
    lu = spla.splu(A, permc_spec = "NATURAL", diag_pivot_thresh=0, options={"SymmetricMode":True})

    # Extract the lower triangular matrix L from LU decomposition
    L = lu.L  # L is a sparse lower triangular matrix
    U = lu.U

    return (L, U)

def write_matrix_to_excel(matrix, N, filename="matrix.xlsx"):
    """
    function we wrote to read/check our big matrices. 
    Saves the matrix to an xlsx file to check if necessary
    """
    from openpyxl import Workbook

    # Create a new Excel workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    
    # Write the matrix to the Excel sheet
    for index, value in np.ndenumerate(matrix):  # Excel rows start at 1
        try:
            sheet.cell(row=index[0]+1, column=index[1]+1, value=value)
        except:
            r, c, d = row_col_dim(index[0], N)
            sheet.cell(row=index[0]+1, column=1, value=str(f'({r},{c},{d})')) #if it is only a 1d matrix
            sheet.cell(row=index[0]+1, column=2, value=value) #if it is only a 1d matrix


    # Save the workbook to the specified file
    workbook.save(filename)
    print(f"Matrix written to {filename}")

def write_sparse_matrix_to_excel(sparse_matrix, N, filename="matrix.xlsx"):
    """
    Write a sparse matrix to an Excel file.
    
    Args:
        sparse_matrix: A sparse matrix in CSR, LIL, or other formats (scipy.sparse).
        filename: The name of the Excel file to save (default: "matrix.xlsx").
        N: Optional, additional parameter for dimensional interpretation.
    """
    from openpyxl import Workbook
    
    # Convert sparse matrix to COO format for easy row/col iteration
    coo_matrix = sparse_matrix.tocoo()
    
    # Create a new Excel workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active
    
    # Write the non-zero elements to the Excel sheet
    for row, col, value in zip(coo_matrix.row, coo_matrix.col, coo_matrix.data):
        try:
            # Write the row, column, and value to the sheet
            sheet.cell(row=row + 1, column=col + 1, value=value)
        except Exception as e:
            if N:
                r, c, d = row_col_dim(row, N)  # Convert index to multi-dimensional coordinates
                sheet.cell(row=row + 1, column=1, value=f"({r},{c},{d})")
                sheet.cell(row=row + 1, column=2, value=value)
            else:
                raise e

    # Save the workbook to the specified file
    workbook.save(filename)
    print(f"Sparse matrix written to {filename}")


@time_it
def forward_substitution(C, f):

    n = C.shape[0]
    y = np.zeros(n)  # Initialize the solution vector with zeros

    for i in range(n):
        # Compute y[i]
        y[i] = (f[i] - np.dot(C[i, :i], y[:i])) / C[i, i]

    return y

@time_it
def forward_substitution_sparse(C, f):
    """
    Solve C * y = f using forward substitution.
    C: Sparse lower triangular matrix in CSR format.
    f: Dense right-hand side vector.
    """

    if not isinstance(C, csr_matrix):
        raise ValueError("Matrix C must be a scipy.sparse.csr_matrix")

    n = C.shape[0]
    y = np.zeros(n)  # Solution vector



    for i in range(n):
        # Extract the row indices and data for nonzero elements in the current row
        row_start = C.indptr[i]
        row_end = C.indptr[i + 1]
        col_indices = C.indices[row_start:row_end]
        values = C.data[row_start:row_end]

        # Compute the dot product for nonzero entries in the current row up to column i
        dot_product = 0
        for idx, col in enumerate(col_indices):
            if col != i:  # Only include entries before the diagonal
                dot_product += values[idx] * y[col]

        # Solve for the current entry in y
        y[i] = (f[i] - dot_product) / values[-1]

    return y

@time_it
def backward_substitution(C, y):
    
    n = C.shape[0]
    u = np.zeros(n)  # Initialize the solution vector with zeros

    for i in range(n - 1, -1, -1):  # Iterate backwards from n-1 to 0
        # Compute y[i]
        u[i] = (y[i] - np.dot(C[i + 1:, i], u[i + 1:])) / C[i, i]

    return u

@time_it
def backward_substitution_sparse(U, y):
    """
    Solve C * u = y using backward substitution.
    C: Sparse upper triangular matrix in CSR format.
    y: Dense right-hand side vector.
    """

    if not isinstance(U, csr_matrix):
        raise ValueError("Matrix C must be a scipy.sparse.csr_matrix")
    
    n = U.shape[0]
    u = np.zeros(n)  # Solution vector

    for i in range(n - 1, -1, -1):  # Iterate backwards from n-1 to 0
        # Extract the row indices and data for nonzero elements in the current row
        row_start = U.indptr[i]
        row_end = U.indptr[i + 1]
        col_indices = U.indices[row_start:row_end]
        values = U.data[row_start:row_end]

        # Compute the dot product for nonzero entries after the diagonal
        dot_product = 0
        for idx, col in enumerate(col_indices):
            if col != i:  # Only include entries after the diagonal
                dot_product += values[idx] * u[col]
            

        # Solve for the current entry in u
        u[i] = (y[i] - dot_product) / values[0]

    return u

def exact_u(
        h, 
        N, 
        d
    ):
    """
    Creates the exact solution 

    3D:
    u(x,y,z) = sin(x*y*z)

    2D:
    u(x,y) = sin(x*y)
    """
    
    u_ex = np.zeros(N**d)
    for p in range(N**d):
        i, j, k = row_col_dim(p, N)
        x, y, z = i*h, j*h, k*h
        if d == 2:
            u_ex[p] = np.sin(x*y)
        else:
            u_ex[p] = np.sin(x*y*z)
    return u_ex

@time_it        
def u_direct_cholesky(
    Ah,
    fh,
    d,
    N
):

    # Create the cholesky decomposition
    C, dt = cholesky_banded(Ah,N)
    y, dt = forward_substitution(
        C,
        fh
    )
    u_ch, dt = backward_substitution(
        C,
        y
    )
    return u_ch

@time_it        
def u_direct_cholesky_sparse(
    Ah_sparse,
    fh,
    N
):
    # calculating C
    L_sparse, U_sparse = sparse_lu(Ah_sparse)

    # calculating y
    y_sparse = forward_substitution_sparse(
        L_sparse.tocsr(),
        fh
    )

    u_lu = backward_substitution_sparse(
        U_sparse.tocsr(),
        y_sparse
    )

    return u_lu

@time_it
def ssor(A, f, u0=None, omega=1.0, tol=1e-8, max_iter=1000):
    """
    Solves the linear system Au = f using the SSOR method.

    """
    n = A.shape[0]
    if u0 is None:
        u = np.zeros(n)
    else:
        u = u0.copy()
    
    D = A.diagonal()
    for iteration in range(max_iter):
        
        # Forward sweep
        for i in range(n):
            sigma = u[i]  # Store previous value
            sum1 = A[i, :i].dot(u[:i])  # Lower triangular part
            sum2 = A[i, i+1:].dot(u[i+1:])  # Upper triangular part
            u[i] = (f[i] - sum1 - sum2) / D[i]
            u[i] = (1 - omega) * sigma + omega * u[i]
        
        # Backward sweep
        for i in range(n-1, -1, -1):
            sigma = u[i]  # Store previous value
            sum1 = A[i, :i].dot(u[:i])  # Lower triangular part
            sum2 = A[i, i+1:].dot(u[i+1:])  # Upper triangular part
            u[i] = (f[i] - sum1 - sum2) / D[i]
            u[i] = (1 - omega) * sigma + omega * u[i]
        
        # Convergence check
        res = norm(f - A @ u)
        if res < tol:
            print(f"Converged after {iteration + 1} iterations with residual {res:.2e}")
            return u, iteration + 1
    
    print(f"Did not converge after {max_iter} iterations. Residual: {res:.2e}")
    return u, max_iter

if __name__ == "__main__":

    # Define a sparse matrix A and vector f
    omega = 1.5  # Relaxation parameter
    
    h = 1/(2**4)
    N = int(1/h+1) # Actually N+1, but this is the size of the one-d matrix, and now N^2 is the size of the 2d matrix etc
    dimensions = 2

    # Create Ah and fh for correct dimensions and h
    Ah_sparse = create_Ah_sparse(h = h, d = dimensions, N=N)
    fh = create_fh(h = h, d = dimensions, N=N)

    u, iterations = ssor(Ah_sparse, fh, omega=omega)



if False:
    I = np.eye(N)
    dimensions = 2
    o = 8
    e_3d = np.zeros(o)
    elapsed_times_3d = np.zeros(o)

    for i in range(o):
        # p = 2*(i+1)
        p = [3,4,5,6,7,8,9,10][i]
        h = 1/(2**p)
        N = int(1/h+1) # Actually N+1, but this is the size of the one-d matrix, and now N^2 is the size of the 2d matrix etc
        I = np.eye(N)

        # Create Ah and fh for correct dimensions and h
        Ah_sparse = create_Ah_sparse(h = h, d = dimensions, N=N)
        fh = create_fh(h = h, d = dimensions, N=N)

        # calculating u_ch
        u_ch_sparse = u_direct_cholesky_sparse(
            Ah_sparse,
            fh,
            N
        )

        # write_sparse_matrix_to_excel(Ah_sparse)

        # C_sparse, dt = cholesky_banded_sparse(Ah_sparse,N)
        
        # Calculate the max error and save the results
        u_ex = exact_u(
            h,
            N,
            dimensions
        )

        # print(f'Ah has {Ah_sparse.count_nonzero()} non-zeros, \n C has {C_sparse.count_nonzero()} non-zeros \n' 
            # f' Resulting in a fraction: {C_sparse.count_nonzero()/Ah_sparse.count_nonzero()}')

        e = np.max(np.abs(u_ex-u_ch_sparse))
        e_3d[i] = e
        elapsed_times_3d[i] = 2

    # print(e)

    import matplotlib.pyplot as plt

    # Replace these with your own values for h (stepsizes) and e (errors)
    h_values_2d = [1/2**i for i in [2,3,4,5,6,7,8][:o]]  # Example: [1/4, 1/8, ..., 1/256]
    e_2d = [4.36737098e-05, 1.31344434e-05, 3.44386074e-06, 8.71427453e-07,2.18518432e-07, 5.46710783e-08]

    # h_values_3d = [1/2**i for i in [2,3,4,5]]
    # e_3d = [0.12473234, 0.33703658, 0.52853519, 0.66766361]

    # Plot the graph
    plt.figure(figsize=(8, 6))
    N_2d = [int(1/h) for h in h_values_2d]
    plt.loglog(N_2d, e_3d, marker='o', label="Error vs Stepsize 2D")
    # plt.loglog(h_values_3d, e_3d, marker='o', label="Error vs Stepsize 3D")

    # Add labels and title
    plt.xlabel("matrix points (N)")
    plt.ylabel("Error")
    plt.title("Log-Log Plot of Stepsize vs Error")
    plt.grid(which="both", linestyle="--", linewidth=0.5)
    plt.legend()

    # Show the plot
    plt.show()
